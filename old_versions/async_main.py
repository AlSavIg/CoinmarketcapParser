import asyncio
import datetime
import json
from typing import Any
from openpyxl import Workbook
import openpyxl
from aiohttp import ClientSession

from scraper.parser.base_classes.config.coinmarket_config import url, headers, params


def create_excel_sheet(titles: tuple):
    _wb = openpyxl.Workbook()
    _sheet = _wb['Sheet']
    row = 1
    for pos, title in enumerate(titles, start=1):
        _sheet.cell(row=row, column=pos).value = title
    return _wb, _sheet


async def get_last_item(session: ClientSession):
    response = await session.get(
        url=url,
        headers=headers,
        params=params)
    return int(json.loads(await response.text())['data']['totalCount'])


async def request_to_data(start: int, session: ClientSession, _sheet):
    params_copy = params.copy()
    params_copy.update(start=start)
    response = await session.get(
        url=url,
        headers=headers,
        params=params_copy
    )
    json_ = json.loads(await response.text())['data']['cryptoCurrencyList']
    main_url = 'https://coinmarketcap.com/currencies/'
    for item in json_:
        name = item['name']
        try:
            rank = item['cmcRank']
            price = item['quotes'][2]['price']  # USD
            ath = item['ath']
            atl = item['atl']
            link = main_url + item['slug'] + '/'
        except KeyError:
            price = ath = atl = rank = link = None
        coefficient = get_coefficient(price, ath, atl, rank)
        write_to_excel(
            (name, rank, coefficient, link),
            rank + 1,
            _sheet
        )


def get_coefficient(price, ath, atl, rank):
    if price == 0:
        return 'PRICE EQUALS ZERO'
    else:
        return (atl / price / rank) if atl > 0 and atl is not None else 1 / rank


def write_to_excel(vals, row, _sheet):
    for i, rec in enumerate(vals, start=1):
        _sheet.cell(row=row, column=i).value = rec


def exec_time_decorator(func):
    async def wrapper(*args, **kwargs):
        start_time = datetime.datetime.now()
        result = await func(*args, **kwargs)
        print(datetime.datetime.now() - start_time)
        return result

    return wrapper


def format_col_width(sheet):
    import string
    for letter in string.ascii_uppercase[:7]:
        sheet.column_dimensions[letter].width = 30


@exec_time_decorator
async def get_data() -> tuple[Workbook, Any, int, str]:
    titles = (
        'name',
        'rank',
        'coef_on_coinmarketcap',
        'link'
    )

    file_name = f'{datetime.datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}_coinmarketcap.xlsx'

    wb, ws = create_excel_sheet(titles=titles)

    async with ClientSession() as session:
        last_item_num = await get_last_item(session)
        step = int(params['limit'])
        tasks = [
            asyncio.create_task(request_to_data(i, session, ws)) for i in range(1, last_item_num, step)
        ]
        await asyncio.gather(*tasks)

    ws.auto_filter.ref = f'A1:E{last_item_num + 1}'
    ws.auto_filter.add_sort_condition(f'C2:C{last_item_num + 1}')

    format_col_width(sheet=ws)
    # _wb.save(filename=file_name)

    return wb, ws, last_item_num, file_name


async def main():
    print(await get_data())


if __name__ == '__main__':
    asyncio.run(main())
